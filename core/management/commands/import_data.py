from django.core.management.base import BaseCommand
import openpyxl
from django.utils.timezone import now
from django.db import transaction
import re
from decimal import Decimal

from core.models import Student, Attendance, Session, Course, Enrollment, Assessment, AssessmentType, Certificate, Statistic

class Command(BaseCommand):
    help = "Импорт данных из Excel"

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Путь к файлу Excel')

    def handle(self, *args, **options):
        # Загрузка файла Excel с поддержкой форматирования для цветов
        wb_colors = openpyxl.load_workbook(options['filepath'], data_only=False)
        ws_colors = wb_colors.active
        
        # Загрузка файла Excel с вычисленными значениями для данных
        wb_data = openpyxl.load_workbook(options['filepath'], data_only=True)
        ws_data = wb_data.active
        
        # Используем ws_colors для цветов и ws_data для значений
        ws = ws_colors  # Основная работа с цветами

        # Словарь для хранения цветов свидетельств и их значений
        # Соответствие согласно легенде в строках 75-85
        cert_colors = {
            'FFFFFF00': Certificate.Status.CONDITIONALLY,     # желтый - Условно-освидетельствованны: не все выполнено
            'FF00B0F0': Certificate.Status.IN_PROGRESS,       # синий - Приготовить свидетельства  
            'FF7030A0': Certificate.Status.CONTROL_RECEIVED,  # фиолетовый - Условно-освидетельствованны: поступила контрольная
            'theme_9': Certificate.Status.COMPLETED           # theme color 9 - Свидетельства готовы в э-форме
        }

        # Функция для извлечения цвета ячейки
        def get_cell_color(cell):
            """Извлекает цвет заливки ячейки"""
            if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                color_obj = cell.fill.fgColor
                # Сначала проверяем theme color
                if hasattr(color_obj, 'type') and color_obj.type == 'theme' and hasattr(color_obj, 'theme'):
                    return f"theme_{color_obj.theme}"
                # Потом проверяем RGB цвет
                elif hasattr(color_obj, 'rgb') and color_obj.rgb:
                    try:
                        rgb = str(color_obj.rgb)
                        if rgb and rgb != "00000000" and rgb != "FFFFFFFF":
                            return rgb
                    except:
                        pass
            return None

        # Проверяем наличие цветов в легенде таблицы
        color_legend = {
            76: "Условно-освидетельствованны: не все выполнено",  # FFFFFF00 (желтый)
            79: "Приготовить свидетельства",                     # FF00B0F0 (синий)
            82: "Условно-освидетельствованны: поступила контрольная",  # FF7030A0 (фиолетовый)
            85: "Свидетельства готовы в э-форме"                 # theme_9
        }
        
        # Выводим информацию о цветах в легенде
        for row, description in color_legend.items():
            cell = ws.cell(row=row, column=2)
            color_code = get_cell_color(cell)
            if color_code:
                self.stdout.write(f"Легенда: '{description}' имеет цвет {color_code}")
            else:
                self.stdout.write(f"Легенда: '{description}' - цвет не найден")
        
        # Список слов, которые указывают, что строка не содержит данных о студенте
        non_student_patterns = [
            r'^\d+-\d+\s*=',  # Например: "91-100 = очень хорошо"
            r'^Условно-освидетельствованны',
            r'^Приготовить свидетельства',
            r'^Свидетельства готовы',
            r'^Приостановленное обучение',
            r'^\d+$'  # Только цифры
        ]

        # Персональная успеваемость не является предметом, а показателями статистики
        stat_columns = {
            'total_courses': None,      # Кол. прослушанных предметов
            'certified': None,          # Кол. освидетельствованных предметов
            'uncertified': None,        # Кол. неосвидетельствованных предметов
            'sessions_missed': None,    # Кол. пропущенных сессий
            'sessions_attended': None,  # К-во сессий с момента начала обучения
            'sessions_late': None       # К обуч. приступил с опозданием на (X) сессий
        }

        # Сохраняем колонки с данными о присутствии для дальнейших расчетов
        presence_columns = {}

        # 1. Парсим структуру сессий
        self.stdout.write("Определение структуры сессий...")
        sessions_data = []
        current_session = None

        # Проходим по всем ячейкам во второй строке для определения сессий и предметов
        for cell in ws[2]:
            if not cell.value or not isinstance(cell.value, str):
                continue
                
            cell_value = cell.value.strip()
            
            # Проверяем, является ли ячейка заголовком сессии (формат: "X с")
            session_match = re.match(r'^(\d+)\s*с.*$', cell_value)
            if session_match:
                session_number = int(session_match.group(1))
                # Создаем или получаем объект сессии
                session_obj, created = Session.objects.get_or_create(
                    session_number=session_number
                )
                if created:
                    self.stdout.write(f"Создана сессия №{session_number}")
                
                # Запоминаем текущую сессию для привязки предметов
                current_session = {
                    'obj': session_obj,
                    'number': session_number,
                    'column': cell.column,
                    'courses': []
                }
                sessions_data.append(current_session)
            # Если это название предмета и у нас есть текущая сессия
            elif current_session and cell_value:
                # Пропускаем "Персональную успеваемость", т.к. это не предмет
                if "Персональная успеваемость" in cell_value:
                    self.stdout.write(f"Обнаружена 'Персональная успеваемость' в колонке {cell.column}, будет обработана как статистика")
                    continue
                    
                course_name = cell_value
                # Создаем или получаем объект курса
                course_obj, created = Course.objects.get_or_create(
                    title=course_name,
                    session=current_session['obj'],
                    defaults={'description': ''}
                )
                if created:
                    self.stdout.write(f"Создан курс '{course_name}' для сессии №{current_session['number']}")
                
                # Запоминаем колонку для привязки данных о курсе
                current_session['courses'].append({
                    'obj': course_obj,
                    'name': course_name,
                    'column': cell.column,
                    'assessment_types': []
                })

        # Отображаем информацию о найденных сессиях и предметах
        for session in sessions_data:
            self.stdout.write(f"Сессия №{session['number']} содержит {len(session['courses'])} предметов:")
            for course in session['courses']:
                self.stdout.write(f"  - {course['name']} (колонка {course['column']})")

        # 2. Парсим типы зачетов и определяем колонки с данными
        self.stdout.write("Парсинг типов зачетов и колонок с данными...")
        
        # Список всех колонок сертификатов для обработки
        all_certificate_columns = []
        
        for cell in ws[3]:  # Третья строка содержит названия типов зачетов
            if not cell.value:
                continue
            
            cell_value = str(cell.value).strip()
            
            # Определяем колонки с присутствием
            if "Присутствие" in cell_value:
                # Определяем, к какой сессии относится эта колонка присутствия
                for session in sessions_data:
                    if abs(session['column'] - cell.column) < 10:  # примерное расстояние
                        session['presence_column'] = cell.column
                        presence_columns[session['number']] = cell.column
                        self.stdout.write(f"Для сессии {session['number']} колонка присутствия: {cell.column}")
            
            # Определяем колонки статистики (персональная успеваемость)
            elif "Кол. прослушаных предметов" in cell_value:
                stat_columns['total_courses'] = cell.column
                self.stdout.write(f"Найдена колонка статистики: 'Кол. прослушаных предметов' в колонке {cell.column}")
            elif "Кол. освидетельствованных предметов" in cell_value:
                stat_columns['certified'] = cell.column
                self.stdout.write(f"Найдена колонка статистики: 'Кол. освидетельствованных предметов' в колонке {cell.column}")
            elif "Кол. неосвидетельствованных предметов" in cell_value:
                stat_columns['uncertified'] = cell.column
                self.stdout.write(f"Найдена колонка статистики: 'Кол. неосвидетельствованных предметов' в колонке {cell.column}")
            elif "Кол. пропущеных сессий" in cell_value:
                stat_columns['sessions_missed'] = cell.column
                self.stdout.write(f"Найдена колонка статистики: 'Кол. пропущеных сессий' в колонке {cell.column}")
            elif "К-во  сессий с момента начала обучения" in cell_value:
                stat_columns['sessions_attended'] = cell.column
                self.stdout.write(f"Найдена колонка статистики: 'К-во сессий с момента начала обучения' в колонке {cell.column}")
            elif "К обуч. приступил с опозданием" in cell_value:
                stat_columns['sessions_late'] = cell.column
                self.stdout.write(f"Найдена колонка статистики: 'К обуч. приступил с опозданием' в колонке {cell.column}")
            
            # Определяем колонки с результатами
            elif "Результат" in cell_value:
                # Находим, к какому курсу относится этот результат
                for session in sessions_data:
                    for course in session['courses']:
                        if abs(course['column'] - cell.column) < 7:  # примерное расстояние
                            course['result_column'] = cell.column
                            self.stdout.write(f"Для курса '{course['name']}' колонка результата: {cell.column}")
            
            # Определяем колонки со свидетельствами
            elif "Свидетельств" in cell_value or "Свидетельст" in cell_value:
                # Добавляем в общий список колонок сертификатов
                all_certificate_columns.append({
                    'column': cell.column,
                    'name': cell_value
                })
                
                # Находим, к какому курсу относится это свидетельство
                course_found = False
                for session in sessions_data:
                    for course in session['courses']:
                        if abs(course['column'] - cell.column) < 7:  # примерное расстояние
                            course['certificate_column'] = cell.column
                            self.stdout.write(f"Для курса '{course['name']}' колонка свидетельства: {cell.column}")
                            course_found = True
                            break
                    if course_found:
                        break
                
                if not course_found:
                    self.stdout.write(f"Найдена независимая колонка сертификатов: {cell_value} (колонка {cell.column})")
            
            # Определяем колонки с типами зачетов
            elif cell_value not in ["Ф.И.О."]:
                # Проверяем, есть ли над ячейкой процент веса
                weight_cell_value = ws.cell(row=2, column=cell.column).value
                weight = None
                if weight_cell_value and isinstance(weight_cell_value, str) and "%" in weight_cell_value:
                    # Парсим процент из строки (например, "75%")
                    weight_match = re.search(r'(\d+)%', weight_cell_value)
                    if weight_match:
                        weight = Decimal(weight_match.group(1)) / Decimal(100)
                        self.stdout.write(f"Найден вес {weight * 100}% для типа зачета '{cell_value}'")
                
                # Создаем или получаем тип зачета
                assessment_type, created = AssessmentType.objects.get_or_create(
                    name=cell_value,
                    defaults={'weight': weight}
                )
                if created:
                    self.stdout.write(f"Создан тип зачета '{cell_value}' с весом {weight * 100 if weight else 'не указан'}%")
                elif weight and assessment_type.weight != weight:
                    assessment_type.weight = weight
                    assessment_type.save()
                    self.stdout.write(f"Обновлен вес для типа зачета '{cell_value}': {weight * 100}%")
                
                # Находим, к какому курсу относится этот тип зачета
                for session in sessions_data:
                    for course in session['courses']:
                        if abs(course['column'] - cell.column) < 7:  # примерное расстояние
                            # Добавляем информацию о колонке с типом зачета
                            course['assessment_types'].append({
                                'obj': assessment_type,
                                'name': cell_value,
                                'column': cell.column,
                                'weight': weight
                            })
                            self.stdout.write(f"Для курса '{course['name']}' тип зачета '{cell_value}' (колонка {cell.column})")
                            break
        
        self.stdout.write(f"Найдено {len(all_certificate_columns)} колонок сертификатов")

        # 3. Импорт данных студентов
        self.stdout.write("Импорт данных студентов...")
        with transaction.atomic():
            # Получаем список студентов
            students = []
            suspended_students = []
            
            # Ищем строку с надписью "Приостановленное обучение"
            suspended_row = None
            for row in range(4, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=2).value
                if cell_value and "Приостановленное обучение" in str(cell_value):
                    suspended_row = row
                    self.stdout.write(f"Найдена строка с надписью 'Приостановленное обучение': {suspended_row}")
                    break
            
            for row in range(4, ws.max_row + 1):
                student_name = ws.cell(row=row, column=2).value
                if not student_name:
                    continue
                
                # Проверяем, является ли строка не-студентом (пояснение, заголовок и т.д.)
                is_non_student = False
                for pattern in non_student_patterns:
                    if re.search(pattern, str(student_name)):
                        is_non_student = True
                        self.stdout.write(f"Пропускаем не-студента: '{student_name}'")
                        break
                
                if is_non_student:
                    continue
                
                # Определяем, является ли студент приостановленным
                is_suspended = suspended_row and row > suspended_row
                student_status = Student.Status.SUSPENDED if is_suspended else Student.Status.ACTIVE
                if is_suspended:
                    self.stdout.write(f"Студент '{student_name}' имеет приостановленное обучение")
                
                # Ищем email в соответствующей колонке (последний столбец)
                email = None
                for col in range(ws.max_column, 1, -1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and '@' in cell_value:
                        email = cell_value
                        break
                
                # Проверяем, есть ли уже студент с таким email
                existing_email_student = None
                if email:
                    try:
                        existing_email_student = Student.objects.get(email=email)
                    except Student.DoesNotExist:
                        pass
                
                # Создаем или получаем студента
                if existing_email_student:
                    # Если нашли студента с таким email, но с другим именем, обновим имя
                    if existing_email_student.full_name != student_name:
                        existing_email_student.full_name = student_name
                        existing_email_student.save()
                        self.stdout.write(f"Обновлено имя для студента с email {email}: {student_name}")
                    
                    # Обновляем статус, если изменился
                    if existing_email_student.status != student_status:
                        existing_email_student.status = student_status
                        existing_email_student.save()
                        self.stdout.write(f"Обновлен статус для студента {student_name} на {student_status}")
                        
                    student = existing_email_student
                    created = False
                else:
                    # Создаем нового студента или получаем по имени
                    student, created = Student.objects.get_or_create(
                        full_name=student_name,
                        defaults={'email': email, 'status': student_status}
                    )
                    if created:
                        self.stdout.write(f"Создан студент '{student_name}' со статусом {student_status}")
                    elif email and not student.email:
                        student.email = email
                        student.save()
                        self.stdout.write(f"Обновлен email для студента '{student_name}'")
                    
                    # Обновляем статус для существующего студента
                    if not created and student.status != student_status:
                        student.status = student_status
                        student.save()
                        self.stdout.write(f"Обновлен статус для существующего студента {student_name} на {student_status}")
                
                students.append((student, row))
                if is_suspended:
                    suspended_students.append((student, row))

            # 4. Импорт данных о посещаемости, оценках и свидетельствах
            self.stdout.write("Импорт данных о посещаемости, оценках и свидетельствах...")
            for student, row in students:
                for session in sessions_data:
                    if 'presence_column' not in session:
                        self.stdout.write(f"Пропускаем сессию {session['number']} - нет колонки присутствия")
                        continue
                    
                    # Проверяем посещаемость
                    presence_value = ws_data.cell(row=row, column=session['presence_column']).value
                    was_present = bool(presence_value)
                    
                    # Для каждого курса в сессии
                    for course in session['courses']:
                        # Создаем запись о зачислении студента на сессию
                        enrollment, _ = Enrollment.objects.get_or_create(
                            student=student,
                            session=session['obj'],
                            defaults={
                                'enrolled_on': now(),
                                'status': Enrollment.Status.COMPLETED if was_present else Enrollment.Status.PLANNED
                            }
                        )
                        
                        # Записываем посещаемость
                        attendance, _ = Attendance.objects.get_or_create(
                            enrollment=enrollment,
                            session=session['obj'],
                            defaults={'present': was_present}
                        )
                        
                        # Для каждого типа зачета в курсе
                        for assessment_type_info in course.get('assessment_types', []):
                            score_value = ws_data.cell(row=row, column=assessment_type_info['column']).value
                            if score_value is not None and isinstance(score_value, (int, float)):
                                # Создаем запись об оценке только если есть реальный балл
                                assessment, _ = Assessment.objects.get_or_create(
                                    enrollment=enrollment,
                                    course=course['obj'],
                                    type=assessment_type_info['obj'],
                                    defaults={
                                        'score': score_value,
                                        'date': now(),
                                        'certificate_issued': False
                                    }
                                )
                        
                        # Если есть колонка с результатом - создаем итоговую оценку
                        if 'result_column' in course:
                            result_value = ws_data.cell(row=row, column=course['result_column']).value
                            if result_value is not None and isinstance(result_value, (int, float)):
                                # Получаем тип оценки "Результат"
                                result_type, _ = AssessmentType.objects.get_or_create(name="Результат")
                                # Создаем или обновляем запись об итоговой оценке
                                result_assessment, _ = Assessment.objects.get_or_create(
                                    enrollment=enrollment,
                                    course=course['obj'],
                                    type=result_type,
                                    defaults={
                                        'score': result_value,
                                        'date': now(),
                                        'certificate_issued': False,
                                        'is_final_grade': True
                                    }
                                )
                        
                        # Обрабатываем сертификат только если есть цветная ячейка
                        if 'certificate_column' in course:
                            cert_cell = ws.cell(row=row, column=course['certificate_column'])
                            cert_value = cert_cell.value
                            
                            # Создаем сертификат только если есть значение И цвет
                            if cert_value:
                                cell_color = get_cell_color(cert_cell)
                                
                                if cell_color and cell_color in cert_colors:
                                    cert_status = cert_colors[cell_color]
                                    
                                    # Ищем связанную итоговую оценку
                                    linked_assessment = None
                                    if 'result_column' in course:
                                        result_type, _ = AssessmentType.objects.get_or_create(name="Результат")
                                        try:
                                            linked_assessment = Assessment.objects.get(
                                                enrollment=enrollment,
                                                course=course['obj'],
                                                type=result_type
                                            )
                                        except Assessment.DoesNotExist:
                                            pass
                                    
                                    # Создаем сертификат
                                    Certificate.objects.update_or_create(
                                        student=student,
                                        course=course['obj'],
                                        defaults={
                                            'assessment': linked_assessment,
                                            'issued_on': now(),
                                            'type': cert_status
                                        }
                                    )
                
                # 5. Расчет и импорт статистики для студента
                self.stdout.write(f"Расчет статистики для студента {student.full_name}")
                
                # Рассчитываем статистику на основе данных из Excel
                total_courses = 0
                certified = 0
                uncertified = 0
                sessions_missed = 0
                sessions_attended = 0
                sessions_late = 0
                
                # Используем данные из таблицы Excel
                if all(col in stat_columns and stat_columns[col] for col in ['total_courses', 'certified', 'uncertified', 'sessions_missed', 'sessions_attended']):
                    # Кол. прослушанных предметов
                    total_val = ws_data.cell(row=row, column=stat_columns['total_courses']).value or 0
                    total_courses = int(total_val) if isinstance(total_val, (int, float)) else 0
                    
                    # Кол. освидетельствованных предметов
                    cert_val = ws_data.cell(row=row, column=stat_columns['certified']).value or 0
                    certified = int(cert_val) if isinstance(cert_val, (int, float)) else 0
                    
                    # Кол. неосвидетельствованных предметов
                    uncert_val = ws_data.cell(row=row, column=stat_columns['uncertified']).value or 0
                    uncertified = int(uncert_val) if isinstance(uncert_val, (int, float)) else 0
                    
                    # Кол. пропущенных сессий
                    missed_val = ws_data.cell(row=row, column=stat_columns['sessions_missed']).value or 0
                    sessions_missed = int(missed_val) if isinstance(missed_val, (int, float)) else 0
                    
                    # К-во сессий с момента начала обучения
                    attended_val = ws_data.cell(row=row, column=stat_columns['sessions_attended']).value or 0
                    sessions_attended = int(attended_val) if isinstance(attended_val, (int, float)) else 0
                    
                    # К обуч. приступил с опозданием на (X) сессий
                    if 'sessions_late' in stat_columns and stat_columns['sessions_late']:
                        late_val = ws_data.cell(row=row, column=stat_columns['sessions_late']).value or 0
                        sessions_late = int(late_val) if isinstance(late_val, (int, float)) else 0
                
                # Создаем или обновляем запись статистики
                Statistic.objects.update_or_create(
                    student=student,
                    defaults={
                        'total_courses': total_courses,
                        'certified': certified,
                        'uncertified': uncertified,
                        'sessions_missed': sessions_missed,
                        'sessions_attended': sessions_attended,
                        'sessions_late': sessions_late
                    }
                )
                self.stdout.write(f"Статистика для {student.full_name}: прослушано {total_courses}, освидетельствовано {certified}, пропущено сессий {sessions_missed}")

        self.stdout.write(self.style.SUCCESS("Импорт данных успешно завершен!"))